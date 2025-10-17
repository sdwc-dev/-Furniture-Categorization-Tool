import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
import re
import openpyxl
from openpyxl.styles import PatternFill

# Configure the page
st.set_page_config(
    page_title="Furniture Categorizer",
    page_icon="🪑",
    layout="wide"
)

# Define category keywords with comprehensive lists
CATEGORY_KEYWORDS = {
    'Loose Furniture': [
        'sofa', 'couch', 'bed', 'table', 'chair', 'desk', 'dresser', 'wardrobe',
        'bookshelf', 'bookcase', 'cabinet', 'console', 'armchair', 'dining',
        'chest', 'bench', 'tv stand', 'entertainment', 'sideboard', 'credenza',
        'armoire', 'nightstand', 'headboard', 'footboard', 'recliner', 'loveseat',
        'sectional', 'ottoman', 'coffee table', 'end table', 'accent table',
        'futon', 'daybed', 'bunk bed', 'bookshelf', 'display cabinet', 'bar cart',
        'filing cabinet', 'office chair', 'dining chair', 'rocking chair'
    ],
    'Outdoor Furniture': [
        'patio', 'outdoor', 'garden', 'deck', 'lawn', 'sun lounger', 'hammock',
        'adironack', 'porch', 'bbq', 'grill', 'beach', 'pool', 'terrace',
        'balcony', 'outdoor sofa', 'patio chair', 'garden bench', 'picnic',
        'camping', 'foldable', 'weatherproof', 'weather-resistant', 'all-weather',
        'rattan', 'teak', 'aluminum', 'resin', 'outdoor dining', 'porch swing',
        'deck chair', 'chaise lounge', 'outdoor cushion', 'gazebo', 'canopy'
    ],
    'Artwork & Accessories': [
        'painting', 'sculpture', 'vase', 'candle', 'frame', 'photo', 'art',
        'decor', 'ornament', 'figurine', 'throw pillow', 'blanket', 'tray',
        'clock', 'mirror', 'wall art', 'print', 'poster', 'tapestry',
        'wall sculpture', 'bowl', 'candle holder', 'candlestick', 'centerpiece',
        'decoration', 'accessory', 'knick knack', 'showpiece', 'art piece',
        'collectible', 'memorabilia', 'pottery', 'ceramic', 'glassware',
        'photo frame', 'picture frame', 'wall clock', 'mantel clock'
    ],
    'Drapery': [
        'curtain', 'drape', 'blind', 'shade', 'valance', 'rod', 'window',
        'drapery', 'sheer', 'blackout', 'voile', 'panel', 'swag', 'cornice',
        'pelmet', 'tieback', 'holdback', 'curtain rod', 'track', 'finial',
        'window treatment', 'roman shade', 'roller blind', 'venetian blind',
        'vertical blind', 'cellular shade', 'pleated shade', 'shutter'
    ],
    'Rug': [
        'rug', 'carpet', 'runner', 'doormat', 'mat', 'kilim', 'persian',
        'oriental', 'area rug', 'floor rug', 'throw rug', 'dhurrie', 'braided',
        'shag', 'wool rug', 'silk rug', 'cotton rug', 'jute', 'sisal', 'seagrass',
        'bamboo', 'needlefelt', 'tufted', 'woven', 'hand-knotted', 'machine made',
        'round rug', 'square rug', 'rectangle rug', 'oval rug'
    ],
    'Lighting': [
        'lamp', 'light', 'chandelier', 'sconce', 'pendant', 'fixture',
        'floor lamp', 'table lamp', 'led', 'bulb', 'ceiling light', 'wall light',
        'track lighting', 'spotlight', 'floodlight', 'downlight', 'uplight',
        'ambient light', 'task light', 'accent light', 'desk lamp', 'reading lamp',
        'bedside lamp', 'night light', 'string light', 'fairy light', 'lantern',
        'torchiere', 'arc lamp', 'banker lamp', 'tiffany lamp', 'crystal',
        'light fixture', 'lamp shade', 'bulb holder'
    ]
}

def advanced_categorization(item_name):
    """
    Advanced categorization using word boundary matching and scoring system
    """
    if pd.isna(item_name) or item_name == '':
        return 'Uncategorized'
    
    item_lower = str(item_name).lower().strip()
    category_scores = {category: 0 for category in CATEGORY_KEYWORDS.keys()}
    
    # Add word boundaries for better matching
    words = re.findall(r'\b\w+\b', item_lower)
    
    for category, keywords in CATEGORY_KEYWORDS.items():
        for keyword in keywords:
            # Check for exact word matches with boundaries
            keyword_pattern = r'\b' + re.escape(keyword) + r'\b'
            if re.search(keyword_pattern, item_lower):
                # Higher score for exact matches
                category_scores[category] += 2
            elif keyword in item_lower:
                # Lower score for partial matches
                category_scores[category] += 1
    
    # Return category with highest score
    best_category = max(category_scores, key=category_scores.get)
    max_score = category_scores[best_category]
    
    # Only return category if score is above threshold
    return best_category if max_score > 0 else 'Uncategorized'

def process_dataframe(df, furniture_column):
    """
    Process the dataframe and add categories while preserving original format
    """
    df_processed = df.copy()
    df_processed['Category'] = df_processed[furniture_column].apply(advanced_categorization)
    
    # Add confidence score (simplified)
    def get_confidence(item_name):
        item_lower = str(item_name).lower()
        category_scores = {category: 0 for category in CATEGORY_KEYWORDS.keys()}
        
        for category, keywords in CATEGORY_KEYWORDS.items():
            for keyword in keywords:
                keyword_pattern = r'\b' + re.escape(keyword) + r'\b'
                if re.search(keyword_pattern, item_lower):
                    category_scores[category] += 2
                elif keyword in item_lower:
                    category_scores[category] += 1
        
        max_score = max(category_scores.values())
        total_possible = 2  # Based on exact match score
        return min(max_score / total_possible, 1.0) if total_possible > 0 else 0.0
    
    df_processed['Confidence'] = df_processed[furniture_column].apply(get_confidence)
    
    return df_processed

def style_uncategorized_rows(row):
    """
    Apply styling to highlight Uncategorized rows
    """
    styles = [''] * len(row)
    if row['Category'] == 'Uncategorized':
        styles = ['background-color: #ffcccc'] * len(row)  # Light red color
    return styles

def to_excel_with_formatting(original_file, df_processed):
    """
    Convert dataframe to Excel while preserving original formatting and merged cells
    """
    output = BytesIO()
    
    # Read the original file with openpyxl to preserve formatting
    workbook = openpyxl.load_workbook(original_file)
    worksheet = workbook.active
    
    # Find the column where we'll add Category and Confidence
    max_col = worksheet.max_column
    
    # Add Category and Confidence headers
    worksheet.cell(row=1, column=max_col + 1, value="Category")
    worksheet.cell(row=1, column=max_col + 2, value="Confidence")
    
    # Create fill for uncategorized rows
    uncategorized_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    
    # Add category and confidence data for each row
    for idx, row_num in enumerate(range(2, worksheet.max_row + 1), start=0):
        if idx < len(df_processed):
            category = df_processed.iloc[idx]['Category']
            confidence = df_processed.iloc[idx]['Confidence']
            
            worksheet.cell(row=row_num, column=max_col + 1, value=category)
            worksheet.cell(row=row_num, column=max_col + 2, value=confidence)
            
            # Apply red background to entire row if uncategorized
            if category == 'Uncategorized':
                for col in range(1, max_col + 3):  # +3 to include new columns
                    worksheet.cell(row=row_num, column=col).fill = uncategorized_fill
    
    # Save the modified workbook
    workbook.save(output)
    return output.getvalue()

# Streamlit UI
def main():
    st.title("🪑 Furniture Categorization Tool")
    st.markdown("Upload your Excel file with furniture names and automatically categorize them!")
    
    # Sidebar for instructions
    with st.sidebar:
        st.header("Instructions")
        st.markdown("""
        1. **Upload** your Excel file (.xlsx, .xls)
        2. **Select** the column containing furniture names
        3. **Process** the data automatically
        4. **Download** the categorized results
        
        **Note:** Original Excel formatting (merged cells, etc.) will be preserved.
        """)
        
        st.header("Categories")
        for category in CATEGORY_KEYWORDS.keys():
            st.write(f"• {category}")
    
    # File upload
    uploaded_file = st.file_uploader(
        "Choose Excel File", 
        type=['xlsx', 'xls'],
        help="Upload your Excel file with furniture product names"
    )
    
    if uploaded_file is not None:
        try:
            # Read the file
            df = pd.read_excel(uploaded_file)
            
            # Display basic info
            st.subheader("📊 Data Information")
            st.write(f"**File:** {uploaded_file.name}")
            st.write(f"**Shape:** {df.shape[0]} rows × {df.shape[1]} columns")
            
            # Column selection
            st.subheader("🔧 Configuration")
            furniture_column = st.selectbox(
                "Select the column containing furniture names:",
                options=df.columns,
                help="Choose the column that contains furniture product names"
            )
            
            if st.button("🚀 Categorize Furniture", type="primary"):
                with st.spinner("Processing your furniture data..."):
                    # Process the data
                    df_processed = process_dataframe(df, furniture_column)
                    
                    # Display results
                    st.subheader("✅ Processing Complete!")
                    
                    # Summary statistics
                    col1, col2, col3, col4 = st.columns(4)
                    with col1:
                        st.metric("Total Items", len(df_processed))
                    with col2:
                        categorized = len(df_processed[df_processed['Category'] != 'Uncategorized'])
                        st.metric("Categorized", categorized)
                    with col3:
                        uncategorized = len(df_processed[df_processed['Category'] == 'Uncategorized'])
                        st.metric("Uncategorized", uncategorized)
                    with col4:
                        avg_confidence = df_processed['Confidence'].mean()
                        st.metric("Avg Confidence", f"{avg_confidence:.1%}")
                    
                    # Category distribution
                    st.subheader("📈 Category Distribution")
                    category_counts = df_processed['Category'].value_counts()
                    st.bar_chart(category_counts)
                    
                    # Display processed data with styling
                    st.subheader("📋 Processed Data")
                    
                    # Apply styling to highlight uncategorized rows
                    styled_df = df_processed.style.apply(style_uncategorized_rows, axis=1)
                    
                    # Also apply gradient to confidence column
                    styled_df = styled_df.background_gradient(
                        subset=['Confidence'], 
                        cmap='YlGnBu'
                    )
                    
                    st.dataframe(
                        styled_df,
                        width='stretch'
                    )
                    
                    # Download section
                    st.subheader("📥 Download Results")
                    
                    st.info("**Preserve Original Format**")
                    st.write("Download the categorized data with all original Excel formatting preserved including merged cells, styles, and layout.")
                    
                    try:
                        excel_data = to_excel_with_formatting(uploaded_file, df_processed)
                        st.download_button(
                            label="📥 Download Categorized Excel File",
                            data=excel_data,
                            file_name="categorized_furniture.xlsx",
                            mime="application/vnd.ms-excel",
                            help="Download with original Excel formatting preserved",
                            type="primary"
                        )
                    except Exception as e:
                        st.error(f"Error preserving original format: {str(e)}")
        
        except Exception as e:
            st.error(f"Error processing file: {str(e)}")
            st.info("Please make sure you've uploaded a valid Excel file and selected the correct column.")

if __name__ == "__main__":
    main()